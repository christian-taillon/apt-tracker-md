import os
import openpyxl
import logging
import re
import sys
import argparse

def sanitize_filename(filename):
    return re.sub(r'[\\/*?:"<>|]', "", filename).strip()

def modify_content(content):
    modification_context = ['Panda', 'Bear', 'Chollima', 'Crane', 'Kitten', 'Tiger', 'Buffalo', 'Hawk', 'Leopard', 'Lynx', 'Wolf', 'Ocelot', 'Sphinx', 'Saiga', 'Spider', 'Jackal', 'Bat']
    
    lines = content.split('\n')
    modified_lines = []
    
    for line in lines:
        words = line.strip().split()
        if len(words) == 2 and words[-1] in modification_context:
            modified_line = f"[[{' '.join(words).upper()}]]"
        else:
            modified_line = line.upper()
        
        modified_lines.append(modified_line)
    
    return '\n'.join(modified_lines)

def modify_toolset_malware(content):
    lines = content.split('\n')
    modified_lines = []
    in_toolset_section = False
    
    for line in lines:
        if line.startswith("## TOOLSET / MALWARE"):
            in_toolset_section = True
            modified_lines.append(line)
        elif in_toolset_section and line.strip():
            tools = line.split(',')
            modified_tools = []
            for tool in tools:
                tool = tool.strip()
                if len(tool.split()) <= 2:
                    tool = f"[[{tool}]]"
                modified_tools.append(tool)
            modified_lines.append(', '.join(modified_tools))
            in_toolset_section = False
        else:
            modified_lines.append(line)
    
    return '\n'.join(modified_lines)

def excel_to_markdown_files(excel_file):
    logging.basicConfig(level=logging.INFO,
                        format='%(asctime)s - %(levelname)s - %(message)s',
                        handlers=[
                            logging.FileHandler('conversion.log'),
                            logging.StreamHandler(sys.stdout)
                        ])

    logging.info(f"Starting conversion of {excel_file}")

    try:
        if not os.path.exists(excel_file):
            raise FileNotFoundError(f"The file {excel_file} does not exist.")

        workbook = openpyxl.load_workbook(excel_file)
        logging.info("Excel file loaded successfully")

        sheets_to_process = ['China', 'Russia', 'North Korea', 'Iran', 'Israel', 'NATO', 'Middle East', 'Others', 'Unknown']

        for sheet_name in sheets_to_process:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                logging.info(f"Processing sheet: {sheet_name}")

                os.makedirs(sheet_name, exist_ok=True)
                logging.info(f"Created directory: {sheet_name}")

                headers = [cell.value for cell in sheet[2] if cell.value is not None]

                for row_index, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
                    if row[0] and row[0] != '?' and str(row[0]).strip():
                        sanitized_name = sanitize_filename(str(row[0]))
                        if sanitized_name:
                            file_name = f"{sanitized_name}.md"
                            file_path = os.path.join(sheet_name, file_name)

                            markdown_content = f"---\norigin: {sheet_name}\n---\n\n"
                            markdown_content += f"[[{row[0]}]]\n\n"
                            for header, value in zip(headers[1:], row[1:]):
                                if value is not None:
                                    markdown_content += f"## {header}\n{value}\n\n"

                            try:
                                modified_content = modify_content(markdown_content)
                                modified_content = modify_toolset_malware(modified_content)
                                
                                with open(file_path, 'w', encoding='utf-8') as file:
                                    file.write(modified_content)
                                logging.info(f"Created file: {file_path}")
                            except IOError as e:
                                logging.error(f"Error writing file {file_path}: {str(e)}")
                        else:
                            logging.warning(f"Skipped row {row_index} in sheet {sheet_name} due to empty sanitized name")
                    else:
                        logging.warning(f"Skipped row {row_index} in sheet {sheet_name} due to missing or invalid main name")
            else:
                logging.warning(f"Sheet {sheet_name} not found in the workbook")

        logging.info("Conversion completed successfully")

    except FileNotFoundError as e:
        logging.error(f"File not found: {str(e)}")
    except Exception as e:
        logging.error(f"An error occurred: {str(e)}", exc_info=True)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Convert APT Groups Excel file to Markdown files.")
    parser.add_argument("-f", "--file", help="Path to the Excel file. If not specified, looks for 'APT Groups and Operations.xlsx' in the current directory.")
    args = parser.parse_args()

    script_dir = os.path.dirname(os.path.abspath(__file__))
    if args.file:
        excel_file_path = args.file
    else:
        excel_file_path = os.path.join(script_dir, 'APT Groups and Operations.xlsx')

    excel_to_markdown_files(excel_file_path)
